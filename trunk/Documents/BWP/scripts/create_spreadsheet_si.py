import unicodecsv
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook
from openpyxl.comments import Comment

spreadsheet_output = "../spreadsheet/SI_FIMS_Spreadsheet_16_Jun_14.xlsx"
bwp_file = "../terms/terms_5_08.tsv"
si_file = "../terms/si_category_order_6_16.tsv"

si_terms = []
si_required = []
with open(si_file, 'rb') as csvfile:
	reader = unicodecsv.DictReader(csvfile, delimiter="\t",encoding='utf-8')
	for row in reader:
		si_terms.append(row['Term Name'])
		if row['Require Level'] == 'required':
			si_required.append(row['Term Name'])

bwp_terms = OrderedDict()
with open(bwp_file, 'rb') as csvfile:
	reader = unicodecsv.DictReader(csvfile, delimiter="\t",encoding='utf-8')
	for row in reader:
		term_dict = {}
		term_name = row['Term Name']
		row.pop('Term Name')
		for key in row:
			if len(row[key]) > 1: #Check to make sure that there is a value in that cell
				term_dict[key] = row[key].encode('ascii','xmlcharrefreplace') #This converts any unicode characters to htmlentities
		bwp_terms[term_name] = term_dict

wb = Workbook()
ws = wb.active
ws.title = "Samples"
for col_id in xrange(1,len(si_terms)+1):
	term = si_terms[col_id-1]
	col = get_column_letter(col_id)
	ws.cell(col+'1').value = term
	if 'Definition' in bwp_terms[term]:
		definition_comment = Comment(bwp_terms[term]['Definition'],'')
		ws.cell(col+'1').comment = definition_comment
	if si_terms[col_id-1] in si_required:
		ws.cell(col+'1').style.font.bold = True
	if (len(si_terms[col_id-1]) > 10):
		ws.column_dimensions[col].width = len(si_terms[col_id-1])+2
	else:
		ws.column_dimensions[col].width = 10

wb.save(filename=spreadsheet_output)

# workbook = xlsxwriter.Workbook(spreadsheet_output)
# worksheet = workbook.add_worksheet('Samples')

# required = workbook.add_format({'bg_color':'#A4C2F4', 'bold':True})
# recommended = workbook.add_format({'bg_color':'#FCE5CD','bold':True})
# optional = workbook.add_format({'bold':True})


# row = 0
# col = 0
# for term in bwp_terms:
# 	if 'FIMS Level' in bwp_terms[term]:
# 		if bwp_terms[term]['FIMS Level'] == 'error':
# 			worksheet.write(row,col,term,required)
# 		else:
# 			worksheet.write(row,col,term,recommended)
# 	else:
# 		worksheet.write(row,col,term,optional)
# 	worksheet.set_column(col,col,len(term)+1)
# 	if 'Definition' in bwp_terms[term]:
# 		worksheet.write_comment(row,col,bwp_terms[term]['Definition'])
# 	col += 1

# workbook.close()


